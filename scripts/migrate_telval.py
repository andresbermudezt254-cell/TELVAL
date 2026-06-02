#!/usr/bin/env python3
"""
migrate_telval.py
Migra el archivo Excel de TELVAL (consolidado_precios_actualizados.xlsx)
a Supabase (tablas: proveedores, productos, proveedor_producto).

Formato real del Excel (hoja "Precios Actualizados"):
  Fila 1-2: Títulos globales
  Fila 3:   Cabecera → Proveedor | Código | Descripción | UM | Vr. Unitario | Fecha
  Fila N:   Header de grupo → col A = "  CODE - NOMBRE" (con espacio inicial), cols B-F vacías
  Fila N+: Productos → col A = "CODE - NOMBRE", B = código, C = descripción, D = UM, E = precio, F = fecha

Uso:
  pip install openpyxl supabase python-dotenv
  python scripts/migrate_telval.py --file "C:/ruta/consolidado_precios_actualizados.xlsx"
  python scripts/migrate_telval.py --file "..." --dry-run
"""

import argparse
import os
import re
import sys
from datetime import datetime, timedelta
from typing import Optional

# Fix encoding for Windows console
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl
    from dotenv import load_dotenv
    from supabase import create_client, Client
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install openpyxl supabase python-dotenv")
    sys.exit(1)

# ─── Config ────────────────────────────────────────────────────────────────────

load_dotenv()

SUPABASE_URL = os.getenv("VITE_SUPABASE_URL") or os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("VITE_SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Set VITE_SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY in .env")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Category inference by keyword
CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "Eléctrico": ["cable", "conductor", "breaker", "interruptor", "transformador", "tablero", "electrico", "eléctric"],
    "Mecánico": ["rodamiento", "eje", "engranaje", "correa", "polea", "resorte", "mecanico", "mecánico"],
    "Hidráulico": ["manguera", "hidraulico", "hidráulic", "cilindro", "bomba", "válvula hidráulic"],
    "Neumático": ["neumatico", "neumático", "compresor", "pistón"],
    "Civil": ["cemento", "concreto", "ladrillo", "hierro", "varilla", "pintura", "civil"],
    "Instrumentación": ["sensor", "transmisor", "encoder", "plc", "modulo", "módulo"],
    "Lubricantes": ["aceite", "grasa", "lubricant"],
    "Seguridad": ["casco", "guante", "boota", "protector", "epp", "seguridad"],
    "Ferretería": ["tornillo", "tuerca", "perno", "remache", "herramienta"],
    "Electrónico": ["tarjeta", "placa", "pcb", "electronico", "electrónico"],
    "Consumibles": ["trapo", "lija", "cinta", "adhesivo", "lubricanste"],
}


def infer_category(name: str) -> str:
    lower = name.lower()
    for cat, kws in CATEGORY_KEYWORDS.items():
        if any(kw in lower for kw in kws):
            return cat
    return "Otros"


def parse_excel_date(value) -> Optional[str]:
    """Convierte fecha de Excel a ISO date string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(value))).date().isoformat()
        except Exception:
            return None
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date().isoformat()
            except ValueError:
                continue
    return None


def parse_supplier_header(cell_value: str) -> tuple[Optional[str], str]:
    """Parsea 'CODIGO - NOMBRE' → (codigo, nombre)."""
    if not cell_value:
        return None, ""
    cleaned = cell_value.strip()
    m = re.match(r'^(\d+)\s*-\s*(.+)$', cleaned)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, cleaned


def migrate(file_path: str, dry_run: bool = False):
    print(f"\nAbriendo: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # La hoja puede llamarse 'Precios Actualizados' u otro nombre
    ws = wb['Precios Actualizados'] if 'Precios Actualizados' in wb.sheetnames else wb.active
    print(f"Hoja activa: {ws.title} | {ws.max_row} filas x {ws.max_column} columnas")

    # Cache para evitar consultas repetidas
    supplier_cache: dict[str, int] = {}   # codigo_interno → id
    product_cache:  dict[str, int] = {}   # codigo_producto → id
    category_cache: dict[str, int] = {}   # nombre_cat → id

    rows_ok = 0
    rows_skip = 0

    # Iterar desde fila 4 (saltar 3 filas de cabecera)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or all(c is None for c in row):
            continue

        col_a = str(row[0] or '').strip()  # Proveedor
        col_b = row[1]                      # Código producto (None en filas de grupo)
        col_c = str(row[2] or '').strip()   # Descripción
        col_d = str(row[3] or 'UND').strip()  # UM
        col_e = row[4]                      # Precio
        col_f = row[5]                      # Fecha

        # Fila de grupo (encabezado de proveedor): col B está vacía
        if col_b is None:
            continue

        # Saltar fila de cabecera de columnas si aparece repetida
        if col_a.lower() in ('proveedor', '') and str(col_b).lower() in ('código', 'codigo', 'ref'):
            continue

        # Fila de producto válida
        if not col_a or not col_b or col_e is None:
            rows_skip += 1
            continue

        # Parsear proveedor desde col A
        prov_codigo, prov_nombre = parse_supplier_header(col_a)
        if not prov_nombre:
            rows_skip += 1
            continue

        # Parsear precio
        try:
            precio = float(str(col_e).replace(',', '.').replace('$', '').replace(' ', ''))
        except (ValueError, TypeError):
            rows_skip += 1
            continue
        if precio <= 0:
            rows_skip += 1
            continue

        prod_codigo = str(col_b).strip()
        prod_nombre = col_c or prod_codigo
        unidad      = col_d or 'UND'
        fecha_iso   = parse_excel_date(col_f)
        cat_nombre  = infer_category(prod_nombre)

        print(f"  [{prov_codigo}] {prov_nombre[:30]:30s} | [{prod_codigo}] {prod_nombre[:40]:40s} | {unidad:4s} | {precio:>10,.0f} | {cat_nombre}")

        if not dry_run:
            # Proveedor
            prov_key = prov_codigo or prov_nombre
            if prov_key not in supplier_cache:
                res = supabase.table("proveedores").select("id").eq("codigo_interno", prov_codigo).execute()
                if res.data:
                    supplier_cache[prov_key] = res.data[0]["id"]
                else:
                    ins = supabase.table("proveedores").insert({
                        "codigo_interno": prov_codigo,
                        "nombre": prov_nombre,
                        "activo": True,
                    }).execute()
                    supplier_cache[prov_key] = ins.data[0]["id"]
            prov_id = supplier_cache[prov_key]

            # Categoría
            if cat_nombre not in category_cache:
                res = supabase.table("categorias").select("id").eq("nombre", cat_nombre).execute()
                if res.data:
                    category_cache[cat_nombre] = res.data[0]["id"]
                else:
                    ins = supabase.table("categorias").insert({"nombre": cat_nombre, "icono": "📦"}).execute()
                    category_cache[cat_nombre] = ins.data[0]["id"]
            cat_id = category_cache[cat_nombre]

            # Producto
            if prod_codigo not in product_cache:
                res = supabase.table("productos").select("id").eq("codigo", prod_codigo).execute()
                if res.data:
                    product_cache[prod_codigo] = res.data[0]["id"]
                else:
                    ins = supabase.table("productos").insert({
                        "codigo": prod_codigo,
                        "nombre": prod_nombre,
                        "unidad_medida": unidad,
                        "categoria_id": cat_id,
                        "activo": True,
                    }).execute()
                    product_cache[prod_codigo] = ins.data[0]["id"]
            prod_id = product_cache[prod_codigo]

            # Precio (upsert)
            supabase.table("proveedor_producto").upsert({
                "proveedor_id":   prov_id,
                "producto_id":    prod_id,
                "precio_unitario": precio,
                "fecha_precio":   fecha_iso,
                "activo":         True,
            }, on_conflict="proveedor_id,producto_id").execute()

        rows_ok += 1

    print(f"\n{'[DRY-RUN] ' if dry_run else ''}✅  {rows_ok} precios procesados, {rows_skip} filas omitidas.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Migrar Excel TELVAL → Supabase")
    parser.add_argument("--file", required=True, help="Ruta al archivo .xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Parsear sin escribir en Supabase")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"ERROR: Archivo no encontrado: {args.file}")
        sys.exit(1)

    migrate(args.file, dry_run=args.dry_run)
